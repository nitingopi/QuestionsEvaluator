import openpyxl
import ast
import textwrap
import re
import sys
import io

def check_answers(file_path):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    column_indices = {header: index for index, header in enumerate(headers)}
    print(f"Column indices: {column_indices}")
    current_question = None
    question_code = []
    print(f"worksheet.iter_rows: {worksheet.iter_rows(min_row=2, values_only=True)}")
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        question_no = row[column_indices['Question No.']]
        print(f"Processing row: {row}")
        print(f"Question No: {question_no}")
        if question_no:  # This is the start of a new question
            # print(f"current_question: {current_question}")
            # print(f"question_code: {question_code}")
            if current_question:
                process_question(current_question, '\n'.join(question_code))
            
            current_question = {
                'no': question_no,
                'level': row[column_indices['Level']],
                'options': [row[column_indices[f'Option{i}']] for i in range(1, 5)],
                'answer': row[column_indices['Answer']]
            }
            question_code = [row[column_indices['Question']]]
            print(f"current_question: {current_question}")
            # question_code = '\n'+question_code[0]
            print(f"question_code: {question_code}")
            process_question(current_question, '\n'.join(question_code))
        else:
            # This row is a continuation of the previous question
            question_code.append(row[column_indices['Question']])

    # Process the last question
    if current_question:
        process_question(current_question, '\n'.join(question_code))

def process_question(question, code_lines):
    print(f"\nProcessing Question {question['no']}:")

    print(f"code_lines: {code_lines}")
    
     # Join all code lines and replace '\n' with actual newlines
    # code = ' '.join(code_lines).replace('\\n', '\n')
    # print(f"Original code:\n{code}")
    
    # Split the code into lines, remove extra spaces from each line, then rejoin
    # lines = code.split('\n')
    # cleaned_lines = [re.sub(r'\s+', ' ', line.strip()) for line in lines]
    # code = '\n'.join(cleaned_lines)
    
    # print(f"Code after cleaning:\n{code}")
    
    # Dedent the code to remove any common leading whitespace
    code = code_lines
    code = textwrap.dedent(code)
    print(f"Code after dedent: {code}")
    print("Code to execute:")
    print(code)
    
    try:
         # Redirect stdout to capture print output
        old_stdout = sys.stdout
        sys.stdout = io.StringIO()
        local_env = {}
        exec(code, {}, local_env)
         # Get the captured output
        output = sys.stdout.getvalue().strip()
        
        # Restore stdout
        sys.stdout = old_stdout
        print(f"Execution result: {output}")
        print(f"Local environment: {local_env}")
        # result = list(local_env.values())[-1]
        # result_str = str(result)
        
        # print(f"Execution result: {result_str}")
        
        matched_option = None
        for i, option in enumerate(question['options'], 1):
            if option and str(option).strip() == output:
                matched_option = i
                break
        
        correct_answer = int(question['answer'])
        if matched_option == correct_answer:
            print(f"Correct! The answer is Option {correct_answer}: {question['options'][correct_answer-1]}")
        else:
            print(f"Incorrect. The expected answer is Option {correct_answer}: {question['options'][correct_answer-1]}")
            print(f"The code output is: {output}")
        
    except Exception as e:
        print(f"Error executing the code: {e}")
        print("Traceback:")
        import traceback
        traceback.print_exc()

# Usage
check_answers('sample.xlsx')